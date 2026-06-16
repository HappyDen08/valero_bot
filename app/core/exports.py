from django.db.models import Sum
from django.http import HttpResponse
from django.utils import timezone


def export_xlsx(queryset, columns, filename):
    """columns: список (заголовок, callable(obj) -> значення)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.append([title for title, _ in columns])
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for obj in queryset:
        ws.append([getter(obj) for _, getter in columns])
    for i, (title, _) in enumerate(columns, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = max(14, len(title) + 4)
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    stamp = timezone.localtime().strftime("%Y-%m-%d_%H%M")
    response["Content-Disposition"] = f'attachment; filename="{filename}_{stamp}.xlsx"'
    wb.save(response)
    return response


PARTICIPANT_COLUMNS = [
    ("ПІБ", lambda p: p.full_name),
    ("Телефон", lambda p: p.phone),
    ("Салон", lambda p: p.salon),
    ("Фабрика", lambda p: p.factory),
    ("Місто", lambda p: p.city),
    ("Telegram ID", lambda p: p.telegram_id),
    ("Квитків", lambda p: p.tickets.count()),
    ("Бонуси, грн", lambda p: float(
        p.sales.filter(status="approved").aggregate(s=Sum("bonus_amount"))["s"] or 0
    )),
    ("Зареєстрований", lambda p: timezone.localtime(p.created_at).strftime("%d.%m.%Y %H:%M")),
]

SALE_COLUMNS = [
    ("№", lambda s: s.id),
    ("Подано", lambda s: timezone.localtime(s.created_at).strftime("%d.%m.%Y %H:%M")),
    ("ПІБ", lambda s: s.participant.full_name),
    ("Салон", lambda s: s.participant.salon),
    ("Фабрика", lambda s: s.participant.factory),
    ("Місто", lambda s: s.participant.city),
    ("Дата продажу", lambda s: s.sale_date.strftime("%d.%m.%Y")),
    ("Виріб", lambda s: s.product_display),
    ("Тканина", lambda s: s.fabric_name),
    ("Бонус, грн", lambda s: float(s.bonus_amount)),
    ("Сума", lambda s: float(s.amount) if s.amount else ""),
    ("Статус", lambda s: s.get_status_display()),
    ("Причина відхилення", lambda s: s.reject_reason),
]

TICKET_COLUMNS = [
    ("Квиток", lambda t: t.code),
    ("ПІБ", lambda t: t.participant.full_name),
    ("Салон", lambda t: t.participant.salon),
    ("Продаж №", lambda t: t.sale_id),
    ("Нараховано", lambda t: timezone.localtime(t.created_at).strftime("%d.%m.%Y %H:%M")),
]
